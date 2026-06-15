#!/usr/bin/env python3
"""
make_tables.py — Build the supplementary tables workbook from pipeline outputs.

Writes tables/Supplementary_Tables.xlsx with one sheet per table:
  S1  BUSCO summary (genome + protein modes)
  S2  Assembly statistics (QUAST)
  S3  Pangenome partition summary
  S4  Per-species pangenome compartment gene counts
  S5  CAFE5 lineage-level summary
  S6  Key gene family copy numbers
  S7  Repeat composition (RepeatMasker .tbl summaries)
  S8  TE-gene proximity by sample and compartment
"""
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
RES = ROOT / "results"
OUT = ROOT / "tables"
OUT.mkdir(exist_ok=True)

SPECIES_ORDER = ["Cx_molestus", "Cx_pallens", "Cx_pipiens", "Cx_quinquefasciatus"]
ALL_SAMPLES = SPECIES_ORDER + ["Cx_tarsalis"]
LABELS = {
    "Cx_molestus": "Cx. molestus",
    "Cx_pallens": "Cx. pallens",
    "Cx_pipiens": "Cx. pipiens",
    "Cx_quinquefasciatus": "Cx. quinquefasciatus",
    "Cx_tarsalis": "Cx. tarsalis",
}

HEADER_FILL = PatternFill("solid", start_color="2C7FB8")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
ITALIC_FONT = Font(name="Arial", size=10, italic=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header(ws, ncols):
    """Apply header styling to row 1 across ncols columns."""
    for j in range(1, ncols + 1):
        c = ws.cell(row=1, column=j)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def autosize(ws, min_w=10, max_w=42):
    """Set column widths from content."""
    for col_cells in ws.columns:
        col = col_cells[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0)
                       for c in col_cells)
        ws.column_dimensions[col].width = max(min_w, min(max_w, max_len + 2))


def write_df(ws, df, title=None, italic_first_col=False):
    """Write a DataFrame at row 1, applying header style and borders."""
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", size=13, bold=True)
        start = 3
    else:
        start = 1
    # Header row
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    # Body rows
    for i, (_, row) in enumerate(df.iterrows(), start=start + 1):
        for j, val in enumerate(row.values, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if italic_first_col and j == 1:
                c.font = ITALIC_FONT
            else:
                c.font = BODY_FONT
            c.alignment = LEFT if j == 1 else CENTER
            c.border = BORDER
    autosize(ws)


# ---------------------------------------------------------------------------
# Build sheets
# ---------------------------------------------------------------------------

def s1_busco():
    rows = []
    for s in ALL_SAMPLES:
        gpath = RES / "busco" / s / "run_diptera_odb10" / "short_summary.json"
        ppath = RES / "busco_proteins" / s / f"short_summary.specific.diptera_odb10.{s}.json"
        row = {"Species": LABELS[s]}
        for mode, p in [("Genome", gpath), ("Protein", ppath)]:
            if not p.exists():
                continue
            j = json.loads(p.read_text())["results"]
            row[f"{mode} Complete %"] = j.get("Complete percentage")
            row[f"{mode} Single %"] = j.get("Single copy percentage")
            row[f"{mode} Duplicated %"] = j.get("Multi copy percentage")
            row[f"{mode} Fragmented %"] = j.get("Fragmented percentage")
            row[f"{mode} Missing %"] = j.get("Missing percentage")
        rows.append(row)
    return pd.DataFrame(rows)


def s2_quast():
    df = pd.read_csv(RES / "quast" / "transposed_report.tsv", sep="\t")
    keep = ["Assembly", "Total length", "# contigs", "Largest contig",
            "N50", "N90", "GC (%)", "# N's per 100 kbp"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df["Total length (Mb)"] = (df["Total length"].astype(int) / 1e6).round(2)
    df["N50 (Mb)"] = (df["N50"].astype(int) / 1e6).round(3)
    df["N90 (Mb)"] = (df["N90"].astype(int) / 1e6).round(3)
    df["Largest contig (Mb)"] = (df["Largest contig"].astype(int) / 1e6).round(2)
    df["Species"] = df["Assembly"].map(LABELS)
    df = df.rename(columns={"# contigs": "Contigs", "# N's per 100 kbp": "Ns / 100 kb"})
    out = df[["Species", "Total length (Mb)", "Contigs", "Largest contig (Mb)",
              "N50 (Mb)", "N90 (Mb)", "GC (%)", "Ns / 100 kb"]].copy()
    # Order: ingroup first, then outgroup
    out["sort"] = out["Species"].apply(
        lambda s: 0 if s != "Cx. tarsalis" else 1)
    out = out.sort_values(["sort", "Species"]).drop(columns="sort").reset_index(drop=True)
    return out


def s3_pangenome():
    df = pd.read_csv(RES / "pangenome" / "pangenome_summary.tsv", sep="\t")
    df = df[df["compartment"] != "outgroup_only"].copy()
    # Add percent column for the top three rows
    main = df[df["compartment"].isin(["core", "shell", "cloud"])]
    total = main["n_orthogroups"].sum()
    df["% of total orthogroups"] = df.apply(
        lambda r: round(r["n_orthogroups"] / total * 100, 2)
        if r["compartment"] in ("core", "shell", "cloud") else "", axis=1)
    df = df.rename(columns={
        "compartment": "Compartment",
        "n_orthogroups": "Orthogroups",
        "total_genes_ingroup": "Total ingroup genes",
    })
    return df


def s4_per_species_partition():
    parts = pd.read_csv(RES / "pangenome" / "partitioned_orthogroups.tsv", sep="\t")
    parts["comp_main"] = parts["compartment"].str.replace(
        r"_Cx_.*_specific$", "", regex=True)
    rows = []
    for s in SPECIES_ORDER:
        row = {"Species": LABELS[s]}
        for comp in ["core", "shell", "cloud"]:
            sub = parts[parts["comp_main"] == comp]
            row[f"{comp.title()} orthogroups (≥1 gene)"] = int((sub[s] > 0).sum())
            row[f"{comp.title()} genes"] = int(sub[s].sum())
        # Species-specific cloud
        ss = parts[parts["compartment"] == f"cloud_{s}_specific"]
        row["Species-specific cloud OGs"] = int(len(ss))
        row["Species-specific cloud genes"] = int(ss[s].sum())
        rows.append(row)
    return pd.DataFrame(rows)


def s5_cafe():
    clade = (RES / "cafe" / "output" / "Gamma_clade_results.txt").read_text()
    rows = []
    for line in clade.strip().splitlines()[1:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        name = re.sub(r"<\d+>$", "", parts[0])
        if name in LABELS:
            exp = int(parts[1]); con = int(parts[2])
            rows.append({"Species": LABELS[name], "Expansions": exp,
                         "Contractions": con, "Net (exp − con)": exp - con})
    gamma = (RES / "cafe" / "output" / "Gamma_results.txt").read_text()
    lam = float(re.search(r"Lambda:\s*([\d\.]+)", gamma).group(1))
    alpha_m = re.search(r"Alpha:\s*([\d\.]+)", gamma)
    alpha = float(alpha_m.group(1)) if alpha_m else None
    sig = sum(1 for _ in open(RES / "cafe" / "significant_families.tsv")) - 1
    df = pd.DataFrame(rows)
    df.attrs["meta"] = {"lambda": lam, "alpha": alpha, "n_sig": sig}
    return df


def s6_key_families():
    df = pd.read_csv(RES / "functional" / "key_families_wide.tsv", sep="\t")
    df = df.rename(columns={"family": "Gene family"})
    rename = {s: LABELS[s] for s in SPECIES_ORDER}
    df = df.rename(columns=rename)
    df = df.set_index("Gene family").loc[
        df.set_index("Gene family").mean(axis=1).sort_values(ascending=False).index
    ].reset_index()
    return df


def parse_rm_tbl(path):
    text = path.read_text()
    total = int(re.search(r"total length:\s+(\d+)\s+bp", text).group(1))
    masked = int(re.search(r"bases masked:\s+(\d+)\s+bp", text).group(1))
    interspersed = int(re.search(
        r"Total interspersed repeats:\s+(\d+)\s+bp", text).group(1))
    sine = int(re.search(r"SINEs:\s+\d+\s+(\d+)\s+bp", text).group(1))
    line = int(re.search(r"^\s*LINEs:\s+\d+\s+(\d+)\s+bp",
                          text, re.MULTILINE).group(1))
    ltr = int(re.search(r"LTR elements:\s+\d+\s+(\d+)\s+bp", text).group(1))
    dna = int(re.search(
        r"DNA transposons\s+\d+\s+(\d+)\s+bp", text).group(1))
    unc_m = re.search(r"Unclassified:\s+\d+\s+(\d+)\s+bp", text)
    unc = int(unc_m.group(1)) if unc_m else 0
    simple = int(re.search(r"Simple repeats:\s+\d+\s+(\d+)\s+bp", text).group(1))
    low = int(re.search(r"Low complexity:\s+\d+\s+(\d+)\s+bp", text).group(1))
    return {
        "Total length (Mb)": round(total / 1e6, 2),
        "Bases masked (%)": round(masked / total * 100, 2),
        "Interspersed TE (%)": round(interspersed / total * 100, 2),
        "  LINEs (%)": round(line / total * 100, 2),
        "  SINEs (%)": round(sine / total * 100, 2),
        "  LTR (%)": round(ltr / total * 100, 2),
        "  DNA TEs (%)": round(dna / total * 100, 2),
        "  Unclassified (%)": round(unc / total * 100, 2),
        "Simple repeats (%)": round(simple / total * 100, 2),
        "Low complexity (%)": round(low / total * 100, 2),
    }


def s7_repeats():
    rows = []
    for s in SPECIES_ORDER:
        tbl = RES / "repeats" / s / f"{s}.fasta.tbl"
        if not tbl.exists():
            continue
        d = parse_rm_tbl(tbl)
        d = {"Species": LABELS[s], **d}
        rows.append(d)
    return pd.DataFrame(rows)


def s8_te_proximity():
    df = pd.read_csv(RES / "repeats" / "te_gene_proximity_summary.tsv", sep="\t")
    df["sample"] = df["sample"].map(LABELS).fillna(df["sample"])
    df["within_1kb"] = (df["within_1kb"] * 100).round(2)
    df = df.rename(columns={
        "sample": "Species",
        "compartment": "Compartment",
        "median": "Median bp",
        "q25": "Q25 bp",
        "q75": "Q75 bp",
        "within_1kb": "Within 1 kb (%)",
        "n": "Genes scored",
    })
    return df


# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    sheets = [
        ("S1_BUSCO", "Table S1. BUSCO completeness (diptera_odb10)", s1_busco(), True),
        ("S2_Assembly_stats", "Table S2. Assembly statistics (QUAST)", s2_quast(), True),
        ("S3_Pangenome_summary", "Table S3. Pangenome partition (orthogroup level)",
         s3_pangenome(), False),
        ("S4_Per_species_partition",
         "Table S4. Pangenome compartment membership by species", s4_per_species_partition(), True),
        ("S5_CAFE_summary", None, s5_cafe(), True),
        ("S6_Key_families",
         "Table S6. Detoxification, chemosensory and immune gene family copy numbers",
         s6_key_families(), False),
        ("S7_Repeat_composition",
         "Table S7. Repeat composition from de novo RepeatModeler + RepeatMasker",
         s7_repeats(), True),
        ("S8_TE_gene_proximity",
         "Table S8. Distance to nearest transposable element per gene, by pangenome compartment",
         s8_te_proximity(), True),
    ]

    for sheet_name, title, df, italic_col1 in sheets:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "S5_CAFE_summary":
            meta = df.attrs["meta"]
            title = (f"Table S5. CAFE5 lineage-level summary "
                     f"(λ = {meta['lambda']:.4f}, α = {meta['alpha']:.3f}, "
                     f"{meta['n_sig']} significant families at p < 0.05)")
        write_df(ws, df, title=title, italic_first_col=italic_col1)
        ws.freeze_panes = "A4" if title else "A2"

    # Front sheet listing contents
    front = wb.create_sheet("Contents", 0)
    front["A1"] = "Supplementary Tables — Cx. pipiens complex pangenome"
    front["A1"].font = Font(name="Arial", size=14, bold=True)
    front["A3"] = "Sheet"
    front["B3"] = "Description"
    for c in (front["A3"], front["B3"]):
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    descriptions = [
        ("S1_BUSCO", "BUSCO completeness for genome- and protein-mode runs (diptera_odb10)"),
        ("S2_Assembly_stats", "QUAST assembly statistics for all five samples"),
        ("S3_Pangenome_summary", "Core / shell / cloud orthogroup counts and species-specific cloud breakdown"),
        ("S4_Per_species_partition", "Orthogroup and gene counts per species in each compartment"),
        ("S5_CAFE_summary", "Per-lineage gene family expansion / contraction counts"),
        ("S6_Key_families", "Copy numbers of P450, OR, GR, CCE, GST, immune, CSP, OBP, IR"),
        ("S7_Repeat_composition", "Genome fraction by repeat class (RepeatMasker)"),
        ("S8_TE_gene_proximity", "Median distance and within-1 kb proximity to nearest TE"),
    ]
    for i, (n, d) in enumerate(descriptions, start=4):
        front.cell(row=i, column=1, value=n).font = BODY_FONT
        front.cell(row=i, column=2, value=d).font = BODY_FONT
        front.cell(row=i, column=1).alignment = LEFT
        front.cell(row=i, column=2).alignment = LEFT
    front.column_dimensions["A"].width = 28
    front.column_dimensions["B"].width = 90

    out_path = OUT / "Supplementary_Tables.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
